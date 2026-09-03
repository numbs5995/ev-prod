# Extract CF-21 data, normalize into entities, emit spec JSON for xlsx_create.py
import openpyxl, json, re

SRC = r"C:/Users/Toyo/Downloads/SPREADSHEET CF-21 (1).xlsx"
wb = openpyxl.load_workbook(SRC, data_only=True, read_only=True)

TALENTS = ["Lyanna", "Noemi", "Deltoriel", "Deltor", "Deltori"]

def detect_talent(name):
    n = str(name)
    low = n.lower()
    for t in TALENTS:
        if t.lower() in low:
            if t == "Deltor": return "Deltoriel"
            return t
    return ""

def base_name(name, talent):
    n = str(name).strip()
    if talent:
        # strip talent name and trailing junk like "(kurangnya)" variants
        for t in TALENTS:
            n = re.sub(re.escape(t), "", n, flags=re.I)
    n = re.sub(r"\(.*?\)", "", n)
    n = re.sub(r"\s+", " ", n).strip(" -–")
    return n

def num(v):
    if v is None: return None
    if isinstance(v, (int, float)): return float(v)
    s = str(v).strip().replace(",", "")
    try: return float(s)
    except: return None

# ---- Items & StockLots from BATCH PRODUCTION ----
bp = wb["BATCH PRODUCTION"]
items, lots = {}, []
next_id = 1
for row in bp.iter_rows(min_row=4, max_row=60, values_only=True):
    row = (tuple(row) + (None,) * 18)[:18]
    name = row[1]
    if not name or not str(name).strip(): continue
    talent = detect_talent(name)
    base = base_name(name, talent)
    key = (base, talent)
    if key not in items:
        items[key] = {"id": f"ITM-{next_id:03d}", "name": base, "talent": talent, "vendor": "", "price": None}
        next_id += 1
    it = items[key]
    lot = {
        "lot_id": f"LOT-{len(lots)+1:03d}", "item_id": it["id"], "item": name.strip(),
        "qty_gacha": num(row[2]), "qty_po": num(row[3]), "qty_ots": num(row[4]),
        "qty_giveaway": num(row[5]), "qty_produced": num(row[6]),
        "status": str(row[8] or "").strip(), "unit_cost": num(row[9]),
        "total_cost": num(row[10]), "pic": str(row[11] or "").strip(),
        "batch": 1, "notes": str(row[7] or "").strip(),
    }
    if any(lot[k] is not None for k in ("qty_gacha","qty_po","qty_ots","qty_giveaway","qty_produced","unit_cost","total_cost")):
        lots.append(lot)

# ---- vendor/price from Merch Production (match by talent+base) ----
mp = wb["Merch Production"]
cur_talent = ""
for row in mp.iter_rows(min_row=4, max_row=120, values_only=True):
    row = (tuple(row) + (None,) * 31)[:31]
    if row[0]: cur_talent = detect_talent(row[0]) or str(row[0]).strip()
    name = row[1]
    if not name: continue
    talent = detect_talent(name) or (cur_talent if cur_talent in TALENTS else "")
    base = base_name(name, talent)
    # normalize aliases
    bn = base.lower()
    if "gantungan kunci" in bn or bn == "keychain": base = "Keychain"
    for it in items.values():
        if it["name"].lower() == base.lower() and it["talent"] == talent:
            v = str(row[5] or "").strip()
            if v and not it["vendor"]: it["vendor"] = v.split("\n")[0].strip()
            c = num(row[8])
            if c and not it.get("mp_cost"): it["mp_cost"] = c

# ---- SaleRecords from DATA HASIL PENJUALAN ----
sl = wb["DATA HASIL PENJUALAN"]
sales = []
for row in sl.iter_rows(min_row=7, max_row=60, values_only=True):
    row = (tuple(row) + (None,) * 28)[:28]
    name = row[1]
    if not name or not str(name).strip(): continue
    talent = detect_talent(name)
    base = base_name(name, talent)
    bn = base.lower()
    if "gantungan kunci" in bn: base = "Keychain"
    sales.append({
        "sale_id": f"SAL-{len(sales)+1:03d}", "item": str(name).strip(), "item_base": base, "talent": talent,
        "qty_sold": num(row[2]), "unit_price": num(row[3]),
        "unit_cost_print": num(row[4]), "unit_cost_pack": num(row[5]),
        "total_sales": num(row[8]), "total_profit": num(row[9]),
        "channel": "OTS",
    })

# match sale cost totals back to items as price reference
it_lookup = {(it["name"].lower(), it["talent"]): it for it in items.values()}
for s in sales:
    it = it_lookup.get((s["item_base"].lower(), s["talent"]))
    if it:
        if s["unit_price"] and not it["price"]: it["price"] = s["unit_price"]
        s["item_id"] = it["id"]
    else:
        s["item_id"] = ""

item_rows = []
for it in items.values():
    item_rows.append({
        "item_id": it["id"], "name": it["name"], "talent": it["talent"], "vendor": it["vendor"],
        "unit_cost_mp": it.get("mp_cost"), "unit_cost_batch": None, "sell_price": it["price"],
    })

spec = {
  "workbook": {"full_calc_on_load": True},
  "sheets": [
    {"name": "README", "rows": [
        ["EV-Prod — normalized view of SPREADSHEET CF-21"],
        [],
        ["Purpose: preview of how EV-Prod would organize the same data. Redundant tabs collapsed into 4 entities."],
        [],
        ["Items", "One row per product+talent variant. Identity master — everything links here."],
        ["StockLots", "One row per production/batch entry with quantities by source (gacha/PO/OTS/giveaway). Entered ONCE."],
        ["CostEntries", "Unit costs per item (from Merch Production + Batch tabs). Sales report pulls from here."],
        ["SaleRecords", "Sales report rows. Profit becomes computed, not retyped."],
        [],
        ["Source file", "SPREADSHEET CF-21 (1).xlsx — tabs: Merch Production, DATA HASIL PENJUALAN, STOCK GACHA CF-21, Harga Packaging/Unit + Printilan, BATCH PRODUCTION"],
        ["Note", "CF-23 data will slot into the same structure."]
    ]},
    {"name": "Items", "rows": [[{"value": h, "type": "header"} if False else h for h in
        ["item_id","name","talent","vendor","unit_cost (MerchProd)","sell_price (from sales)"]]] +
        [[i["item_id"], i["name"], i["talent"], i["vendor"], i["unit_cost_mp"], i["sell_price"]] for i in item_rows],
     "freeze_panes": "A2", "column_widths": [10, 26, 12, 18, 20, 20]},
    {"name": "StockLots", "rows": [["lot_id","item_id","item (original name)","qty_gacha","qty_po","qty_ots","qty_giveaway","qty_produced","status","unit_cost","total_cost","pic","batch","notes"]] +
        [[l["lot_id"],l["item_id"],l["item"],l["qty_gacha"],l["qty_po"],l["qty_ots"],l["qty_giveaway"],l["qty_produced"],l["status"],l["unit_cost"],l["total_cost"],l["pic"],l["batch"],l["notes"]] for l in lots],
     "freeze_panes": "C2", "column_widths": [9,9,28,10,9,9,11,13,12,11,12,9,7,24]},
    {"name": "SaleRecords", "rows": [["sale_id","item_id","item (original)","talent","qty_sold","unit_price","unit_cost_print","unit_cost_pack","total_sales","total_profit","channel"]] +
        [[s["sale_id"],s["item_id"],s["item"],s["talent"],s["qty_sold"],s["unit_price"],s["unit_cost_print"],s["unit_cost_pack"],s["total_sales"],s["total_profit"],s["channel"]] for s in sales],
     "freeze_panes": "C2", "column_widths": [9,9,30,11,9,11,14,14,12,12,9]},
  ]
}

with open(r"C:/Users/Toyo/ev-prod/analysis/spec.json", "w", encoding="utf-8") as f:
    json.dump(spec, f, ensure_ascii=False)
print("items:", len(item_rows), "lots:", len(lots), "sales:", len(sales))
