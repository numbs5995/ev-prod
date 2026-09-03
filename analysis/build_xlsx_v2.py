# Build v2 normalized xlsx: channel column, clean names, Gacha sheet, summary
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

d = json.load(open(r"C:/Users/Toyo/ev-prod/analysis/normalized.json", encoding="utf-8"))
wb = Workbook(); wb.remove(wb.active)
hf = Font(bold=True, color="FFFFFF"); hfill = PatternFill("solid", fgColor="4472C4")
tf = Font(bold=True, size=13)

def add(name, header, rows, widths, title=None):
    ws = wb.create_sheet(name)
    r0 = 1
    if title:
        ws.cell(1, 1, title).font = tf; r0 = 3
    for c, h in enumerate(header, 1):
        cell = ws.cell(r0, c, h); cell.font = hf; cell.fill = hfill
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    for r, row in enumerate(rows, r0 + 1):
        for c, v in enumerate(row, 1):
            ws.cell(r, c, v)
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = f"A{r0+1}"
    return ws

# Summary sheet
sum_rows = [
    ["Items (one per product + talent)", len(d["items"])],
    ["StockLots (batch production entries, pre-event prep)", len(d["lots"])],
    ["SaleRecords (post-event sales, per channel)", len(d["sales"])],
    ["Gacha prizes (from STOCK GACHA CF-21)", len(d["gacha"])],
    [],
    ["Channel legend:"],
    ["PO", "Pre-order sales"],
    ["OTS", "On-the-spot sales at event"],
    ["Auction", "Auctioned items"],
    ["Staff", "Internal buy at production cost, or leftover stock"],
    ["Gacha", "Gacha summary row"],
    [],
    ["Naming fixed:", 'Standee Akrilik = Standee Acrylic; Gantungan Kunci = Keychain; "PO" suffix moved to Channel column'],
]
add("Summary", ["Sheet", "Count / note"], sum_rows, [52, 70], title="EV-Prod — normalized CF-21 data (v2)")

add("Items", ["item_id", "item", "talent"], d["items"], [10, 30, 16],
    title="Item catalog — product × talent. Everything links by item_id.")
add("StockLots", ["lot_id", "item_id", "item", "talent", "qty_gacha", "qty_po", "qty_ots", "qty_giveaway",
                  "qty_produced", "status", "unit_cost", "total_cost", "pic", "notes", "raw_name (source)"],
    d["lots"], [9, 9, 24, 11, 10, 8, 8, 11, 12, 12, 11, 12, 9, 20, 26],
    title="Production batches (prepared before CF-21). Quantities entered ONCE per source.")
add("SaleRecords", ["sale_id", "item_id", "item", "talent", "channel", "qty_sold", "unit_price",
                    "unit_cost_print", "unit_cost_pack", "total_sales", "total_profit", "raw_name (source)"],
    d["sales"], [9, 9, 26, 11, 10, 9, 11, 13, 13, 12, 12, 30],
    title="Post-event sales. Channel = PO / OTS / Auction / Staff / Gacha.")
add("Gacha", ["prize", "qty_total", "alloc Lyanna", "alloc Noemi", "alloc Deltoriel", "alloc other",
              "cost/pcs", "revenue/pcs", "total_cost", "total_revenue", "margin"],
    d["gacha"] + [[]] + d["gacha_meta"], [26, 10, 12, 12, 12, 10, 11, 12, 12, 13, 12],
    title="Gacha pool (STOCK GACHA CF-21). Prizes are items; qty entered here = gacha stock.")

wb.save(r"C:/Users/Toyo/ev-prod/analysis/ev-prod-normalized-preview-v2.xlsx")
print("saved v2")
