# Build normalized preview xlsx directly
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

spec = json.load(open(r"C:/Users/Toyo/ev-prod/analysis/spec.json", encoding="utf-8"))
wb = Workbook()
wb.remove(wb.active)
hdr_font = Font(bold=True)
hdr_fill = PatternFill("solid", fgColor="DDEBF7")

for sh in spec["sheets"]:
    ws = wb.create_sheet(sh["name"])
    for r, row in enumerate(sh["rows"], 1):
        for c, val in enumerate(row, 1):
            cell = ws.cell(row=r, column=c)
            if isinstance(val, dict):
                cell.value = val.get("value")
                if val.get("type") == "date":
                    from datetime import datetime
                    try: cell.value = datetime.fromisoformat(val["value"])
                    except: pass
            else:
                cell.value = val
            if r == 1 or (sh["name"] == "README"):
                pass
    # style first row as header when it looks like one
    first = [c.value for c in ws[1] if c.value]
    if sh["name"] != "README" and first and all(isinstance(v, str) for v in first):
        for c in ws[1]:
            if c.value is not None:
                c.font = hdr_font; c.fill = hdr_fill
                c.alignment = Alignment(vertical="center")
        ws.freeze_panes = "A2"
    for i, w in enumerate(sh.get("column_widths", []), 1):
        ws.column_dimensions[get_column_letter(i)].width = w

wb.save(r"C:/Users/Toyo/ev-prod/analysis/ev-prod-normalized-preview.xlsx")
print("saved")
