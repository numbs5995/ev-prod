# v2: normalized CF-21 with channel labels, normalized names, gacha sheet
import openpyxl, json, re

SRC = r"C:/Users/Toyo/Downloads/SPREADSHEET CF-21 (1).xlsx"
wb = openpyxl.load_workbook(SRC, data_only=True, read_only=True)

TALENTS = ["Lyanna", "Noemi", "Deltoriel"]

def detect_talent(s):
    low = str(s).lower()
    for t in TALENTS:
        if t.lower() in low: return t
    if "deltor" in low: return "Deltoriel"
    return ""

NAME_MAP = [
    (r"standee\s*1:1", "Standee 1:1"),
    (r"standee\s*akrilik|standee\s*acrylic", "Standee Acrylic"),
    (r"gantungan\s*kunci|keychain|ganci\s*lama", "Keychain (old)"),
    (r"ganci\s*baru", "Keychain (new)"),
    (r"lenticular", "Lenticular Card"),
    (r"gantungan", "Keychain (gacha prize)"),
    (r"poster", "Poster A3+"),
    (r"bottle\s*cap\s*pin", "Bottle Cap Pin"),
    (r"lanyard\s*\+\s*card\s*holder|lanyard\s*\+card\s*holder", "Lanyard + Card Holder"),
    (r"lanyard", "Lanyard"),
    (r"sticker\s*pack\s*truck|sticker\s*pack", "Sticker Pack Truck"),
    (r"dakimakura", "Dakimakura"),
    (r"deskmat", "Deskmat"),
    (r"enamel\s*mug", "Enamel Mug"),
    (r"phone\s*case", "Phone Case"),
    (r"emoney", "Emoney"),
    (r"poster\s*a3\+?", "Poster A3+"),
    (r"canvas\s*art", "Canvas Art 40x30cm"),
    (r"card\s*holder", "Card Holder"),
]

def parse_sale_name(raw):
    """Return (item, talent, channel). Channel: PO/OTS/Auction/Staff/Gacha."""
    s = str(raw).strip()
    channel = "OTS"
    low = s.lower()
    if "staff" in low: channel = "Staff"
    elif "auction" in low: channel = "Auction"
    elif "po only" in low or low.endswith(" po") or " po)" in low or " po " in low or "(po" in low: channel = "PO"
    if "gacha" == low: return ("Gacha (summary)", "", "Gacha")
    s2 = re.sub(r"\((po only|auction|kurangnya)\)", "", s, flags=re.I)
    s2 = re.sub(r"\bstaff\b|\bauktion\b|po only", "", s2, flags=re.I)
    s2 = re.sub(r"\bpo\b", "", s2, flags=re.I)
    s2 = re.sub(r"\+\s*parfum\s*\+\s*box khusus", "", s2, flags=re.I).strip(" (+")
    talent = detect_talent(s2)
    item = s2
    for pat, name in NAME_MAP:
        if re.search(pat, s2, flags=re.I):
            item = name; break
    return (item, talent, channel)

def num(v):
    if v is None: return None
    if isinstance(v, (int, float)): return float(v)
    try: return float(str(v).strip().replace(",", ""))
    except: return None

# ---------- Sales (source of item catalog: every sold variant) ----------
sl = wb["DATA HASIL PENJUALAN"]
sales, items, item_ix = [], {}, {}
def get_item(item, talent):
    key = (item, talent)
    if key not in items:
        items[key] = {"item_id": f"ITM-{len(items)+1:03d}", "item": item, "talent": talent}
    return items[key]

for row in sl.iter_rows(min_row=7, max_row=60, values_only=True):
    row = (tuple(row) + (None,)*28)[:28]
    if not row[1] or not str(row[1]).strip(): continue
    item, talent, channel = parse_sale_name(row[1])
    it = get_item(item, talent)
    sales.append({
        "sale_id": f"SAL-{len(sales)+1:03d}", "item_id": it["item_id"],
        "item": item, "talent": talent, "channel": channel,
        "qty_sold": num(row[2]), "unit_price": num(row[3]),
        "unit_cost_print": num(row[4]), "unit_cost_pack": num(row[5]),
        "total_sales": num(row[8]), "total_profit": num(row[9]),
        "raw_name": str(row[1]).strip(),
    })

# ---------- StockLots from BATCH PRODUCTION (prep data, pre-event) ----------
bp = wb["BATCH PRODUCTION"]
lots = []
for row in bp.iter_rows(min_row=4, max_row=60, values_only=True):
    row = (tuple(row) + (None,)*18)[:18]
    if not row[1] or not str(row[1]).strip(): continue
    raw = str(row[1]).strip()
    item, talent, channel = parse_sale_name(raw)
    it = get_item(item, talent)
    lots.append({
        "lot_id": f"LOT-{len(lots)+1:03d}", "item_id": it["item_id"],
        "item": item, "talent": talent,
        "qty_gacha": num(row[2]), "qty_po": num(row[3]), "qty_ots": num(row[4]),
        "qty_giveaway": num(row[5]), "qty_produced": num(row[6]),
        "status": str(row[8] or "").strip(), "unit_cost": num(row[9]),
        "total_cost": num(row[10]), "pic": str(row[11] or "").strip(),
        "batch": 1, "notes": str(row[7] or "").strip(), "raw_name": raw,
    })

# ---------- Gacha sheet from STOCK GACHA CF-21 ----------
gz = wb["STOCK GACHA CF-21"]
gacha = []
for row in gz.iter_rows(min_row=4, max_row=14, values_only=True):
    row = (tuple(row) + (None,)*15)[:15]
    name = row[2]
    if not name: continue
    gacha.append({
        "prize": str(name).strip(), "qty_total": num(row[3]),
        "alloc_lyanna": num(row[4]), "alloc_noemi": num(row[5]),
        "alloc_deltoriel": num(row[6]), "alloc_other": num(row[7]),
        "cost_per_pcs": num(row[9]), "revenue_per_pcs": num(row[10]),
        "total_cost": num(row[11]), "total_revenue": num(row[12]), "margin": num(row[13]),
    })
# gacha rate + easter egg tiers
rate = {}
for row in gz.iter_rows(min_row=18, max_row=21, values_only=True):
    if row[3] is not None:
        rate.setdefault("rates", []).append(round(float(row[3]), 4))
egg = {}
for row in gz.iter_rows(min_row=25, max_row=27, values_only=True):
    if row[3]: egg[str(row[3]).strip()] = num(row[4])

def sheet(name, header, rows):
    return {"name": name, "header": header, "rows": rows}

item_rows = [[it["item_id"], it["item"], it["talent"] or "(none / shared)"] for it in items.values()]
lot_rows = [[l["lot_id"], l["item_id"], l["item"], l["talent"], l["qty_gacha"], l["qty_po"], l["qty_ots"],
             l["qty_giveaway"], l["qty_produced"], l["status"], l["unit_cost"], l["total_cost"], l["pic"], l["notes"], l["raw_name"]] for l in lots]
sale_rows = [[s["sale_id"], s["item_id"], s["item"], s["talent"], s["channel"], s["qty_sold"], s["unit_price"],
              s["unit_cost_print"], s["unit_cost_pack"], s["total_sales"], s["total_profit"], s["raw_name"]] for s in sales]
gacha_rows = [[g["prize"], g["qty_total"], g["alloc_lyanna"], g["alloc_noemi"], g["alloc_deltoriel"], g["alloc_other"],
               g["cost_per_pcs"], g["revenue_per_pcs"], g["total_cost"], g["total_revenue"], g["margin"]] for g in gacha]
egg_rows = [["Gacha prize rate (zonk/mid/gacor):"] + [str(r) for r in rate.get("rates", [])]]
egg_rows += [[k, v] for k, v in egg.items()]

out = {
    "items": item_rows, "lots": lot_rows, "sales": sale_rows,
    "gacha": gacha_rows, "gacha_meta": egg_rows,
    "counts": {"items": len(items), "lots": len(lots), "sales": len(sales), "gacha": len(gacha)},
}
with open(r"C:/Users/Toyo/ev-prod/analysis/normalized.json", "w", encoding="utf-8") as f:
    json.dump(out, f, ensure_ascii=False, indent=1)
print(out["counts"])
